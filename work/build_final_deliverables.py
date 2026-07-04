from __future__ import annotations

import math
import os
import re
import subprocess
from copy import copy
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter, Transformation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


ROOT = Path(os.environ.get("LOGBOOK_ROOT", Path(__file__).resolve().parents[1])).resolve()
OUT = ROOT / "outputs"
WORK = ROOT / "work"
PILOTLOG_EXPORT_XLSX = WORK / "PILOTLOG_export.xlsx"
SYNCED_AUTHORITATIVE_XLSX = WORK / "log_filled_authoritative_synced.xlsx"

XLSX_OUT = OUT / "log filled.xlsx"
A4_OUT = OUT / "ICAO_EASA_A4_landscape_logbook.pdf"
BOOKLET_OUT = OUT / "A5_booklet_A4_portrait_duplex.pdf"

NAVY = colors.HexColor("#1B3A6B")
GOLD = colors.HexColor("#B8962E")
LIGHT_GOLD = colors.HexColor("#F4E8C4")
GRID = colors.HexColor("#BFC6D1")
TEXT = colors.HexColor("#1A1D21")

START = {
    "blk": 13135 * 60 + 59,
    "ngt": 4785 * 60 + 36,
    "ifr": 9637 * 60 + 53,
    "to": 1726,
    "ldg": 1732,
    "pic": 6146 * 60 + 26,
}

ALL_DCS = ["2C1", "3C1", "3C2", "3CX", "4C1", "4C2", "4C3", "4CR", "2CR", "2CO", "4CX"]
PIC_FACTORS = {
    "2C1": 1.0,
    "3C1": 2 / 3,
    "3C2": 1 / 3,
    "3CX": 0.0,
    "4C1": 0.5,
    "4C2": 0.5,
    "4C3": 0.0,
    "4CR": 0.0,
    "2CR": 0.0,
    "2CO": 0.0,
    "4CX": 0.0,
}


@dataclass
class Flight:
    date: datetime
    flt: str
    dep: str
    arr: str
    reg: str
    dc: str
    off: time | None
    on: time | None
    blk_min: int
    takeoff: time | None
    landing: time | None
    act_min: int
    ngt_min: int
    ifr_min: int
    to_count: int
    ldg_count: int
    autoland: str
    toga: str
    pic_min: int
    remarks: str


def minutes(value) -> float:
    if value is None:
        return 0
    if isinstance(value, timedelta):
        return value.total_seconds() / 60
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)):
        return float(value) * 24 * 60
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return 0
        try:
            # Google Sheets exports some duration cells as decimal day strings.
            return float(text) * 24 * 60
        except ValueError:
            pass
        text = text.replace("오전", "").replace("오후", "").strip()
        match = re.match(r"^(\d+):(\d{1,2})(?::(\d{1,2}))?$", text)
        if match:
            return int(match.group(1)) * 60 + int(match.group(2)) + (int(match.group(3) or 0) / 60)
    return 0.0


def cell_time(value) -> time | None:
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, (int, float)):
        total_seconds = int(round((float(value) % 1) * 86400))
        return time((total_seconds // 3600) % 24, (total_seconds % 3600) // 60, total_seconds % 60)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        pm = "오후" in text
        text = text.replace("오전", "").replace("오후", "").strip()
        match = re.match(r"^(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?$", text)
        if match:
            hour = int(match.group(1))
            if pm and hour < 12:
                hour += 12
            return time(hour % 24, int(match.group(2)), int(match.group(3) or 0))
    return None


def duration_fraction(mins: int | float) -> float:
    return float(mins) / 1440.0


def excel_time_fraction(value: time | None) -> float | None:
    if value is None:
        return None
    return (value.hour * 3600 + value.minute * 60 + value.second) / 86400.0


def fmt_hmm(value) -> str:
    mins = int(round(minutes(value) if not isinstance(value, (int, float)) else value))
    return f"{mins // 60}:{mins % 60:02d}"


def fmt_time(value: time | None) -> str:
    return "" if not isinstance(value, time) else f"{value.hour:02d}:{value.minute:02d}"


def pdf_remarks(value: str) -> str:
    text = "" if value is None else str(value)
    return text if text.isascii() else ""


def read_pilotlog_export_flights() -> list[Flight]:
    wb = load_workbook(PILOTLOG_EXPORT_XLSX, data_only=True)
    ws = wb["flt_log"]
    flights: list[Flight] = []
    for r in range(1, ws.max_row + 1):
        date = parse_date(ws.cell(r, 3).value)
        if not date:
            continue
        flt = norm_flight_number(ws.cell(r, 4).value)
        dep = norm_airport(ws.cell(r, 5).value)
        arr = norm_airport(ws.cell(r, 6).value)
        reg = norm_reg(ws.cell(r, 7).value)
        blk_min = minutes(ws.cell(r, 11).value)
        if not (flt and dep and arr and reg):
            continue
        dc = str(ws.cell(r, 8).value or "").strip().upper()
        pic_min = minutes(ws.cell(r, 21).value)
        if pic_min == 0 and blk_min:
            pic_min = blk_min * PIC_FACTORS.get(dc, 0.0)
        flights.append(
            Flight(
                date=date,
                flt=flt,
                dep=dep,
                arr=arr,
                reg=reg,
                dc=dc,
                off=cell_time(ws.cell(r, 9).value),
                on=cell_time(ws.cell(r, 10).value),
                blk_min=blk_min,
                takeoff=cell_time(ws.cell(r, 12).value),
                landing=cell_time(ws.cell(r, 13).value),
                act_min=minutes(ws.cell(r, 14).value),
                ngt_min=minutes(ws.cell(r, 15).value),
                ifr_min=minutes(ws.cell(r, 16).value),
                to_count=int(ws.cell(r, 17).value or 0),
                ldg_count=int(ws.cell(r, 18).value or 0),
                autoland=str(ws.cell(r, 19).value or ""),
                toga=str(ws.cell(r, 20).value or ""),
                pic_min=pic_min,
                remarks=str(ws.cell(r, 28).value or ""),
            )
        )
    flights.sort(key=lambda f: (f.date, f.off or time(0, 0), f.flt))
    return flights


def read_authoritative_source_flights() -> list[Flight]:
    if SYNCED_AUTHORITATIVE_XLSX.exists():
        return read_filled_logbook_flights(SYNCED_AUTHORITATIVE_XLSX)
    print(f"WARNING_SYNCED_AUTHORITATIVE_SOURCE_MISSING {SYNCED_AUTHORITATIVE_XLSX}")
    print(f"WARNING_FALLING_BACK_TO_PILOTLOG_EXPORT {PILOTLOG_EXPORT_XLSX}")
    return read_pilotlog_export_flights()


def normalize_header(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower().replace("$", ""))


def parse_counter_date(value) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d%b%y", "%d.%b.%y", "%Y.%m.%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    nums = [int(x) for x in re.findall(r"\d+", text)]
    if len(nums) >= 3:
        if nums[0] > 1900:
            return datetime(nums[0], nums[1], nums[2])
        return datetime(2000 + nums[2] if nums[2] < 100 else nums[2], nums[1], nums[0])
    return None


def parse_counter_time(value) -> time | None:
    text = str(value or "").strip()
    if not text:
        return None
    text = re.sub(r"([+-]\d+)$", "", text)
    text = {"Jan.00": "0100", "Jan.10": "0110", "Jan.20": "0120", "Jan.24": "0124", "Jan.31": "0131", "Jan.33": "0133", "Jan.37": "0137", "Jan.44": "0144", "Jan.48": "0148", "Jan.50": "0150", "Jan.51": "0151", "Jan.59": "0159"}.get(text, text)
    digits = re.sub(r"\D", "", text)
    if not digits:
        return None
    if len(digits) <= 2:
        hour, minute = 0, int(digits)
    else:
        digits = digits.zfill(4)[-4:]
        hour, minute = int(digits[:2]), int(digits[2:])
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return time(hour, minute)
    return None


def has_previous_day_suffix(value) -> bool:
    return bool(re.search(r"-1\s*$", str(value or "").strip()))


def parse_duration_minutes(value) -> int:
    text = str(value or "").strip()
    match = re.match(r"^(\d+):(\d{1,2})$", text)
    if not match:
        return 0
    return int(match.group(1)) * 60 + int(match.group(2))


def elapsed_minutes(start: time | None, finish: time | None) -> int:
    if not start or not finish:
        return 0
    value = (finish.hour * 60 + finish.minute) - (start.hour * 60 + start.minute)
    if value < 0:
        value += 24 * 60
    return value


def read_roster_lookup() -> dict[tuple[date, str, str, str], tuple[str, str]]:
    wb = load_workbook(PILOTLOG_EXPORT_XLSX, data_only=True)
    ws = wb["Roster1"]
    lookup: dict[tuple[date, str, str, str], tuple[str, str]] = {}
    current_date: datetime | None = None
    for r in range(2, ws.max_row + 1):
        parsed = parse_date(ws.cell(r, 1).value)
        if parsed:
            current_date = parsed
        if not current_date:
            continue
        flt = norm_flight_number(ws.cell(r, 5).value)
        dep = norm_airport(ws.cell(r, 7).value)
        arr = norm_airport(ws.cell(r, 10).value)
        if not (flt and dep and arr):
            continue
        key = (current_date.date(), flt, dep, arr)
        lookup[key] = (str(ws.cell(r, 14).value or ""), infer_csv_dc(flt))
    return lookup


def read_flt_log_lookup() -> dict[tuple[date, str, str, str], tuple[str, str]]:
    wb = load_workbook(PILOTLOG_EXPORT_XLSX, data_only=True)
    ws = wb["flt_log"]
    lookup: dict[tuple[date, str, str, str], tuple[str, str]] = {}
    for r in range(1, ws.max_row + 1):
        parsed = parse_date(ws.cell(r, 3).value)
        if not parsed:
            continue
        flt = norm_flight_number(ws.cell(r, 4).value)
        dep = norm_airport(ws.cell(r, 5).value)
        arr = norm_airport(ws.cell(r, 6).value)
        if flt and dep and arr:
            lookup[(parsed.date(), flt, dep, arr)] = (str(ws.cell(r, 7).value or ""), str(ws.cell(r, 8).value or ""))
    return lookup


def infer_csv_dc(flt: str) -> str:
    return {
        "151": "3C2",
        "152": "3C1",
        "135": "4C2",
        "136": "4C2",
        "801": "2C1",
        "802": "2C1",
        "131": "4C1",
        "132": "4C2",
    }.get(str(flt), "")


def flight_identity(f: Flight) -> tuple:
    return (f.date.date(), f.flt, f.dep, f.arr, f.reg)


def flight_fingerprint(f: Flight) -> tuple:
    return (
        f.dc,
        f.off,
        f.on,
        round(f.blk_min),
        f.takeoff,
        f.landing,
        round(f.act_min),
        round(f.ngt_min),
        round(f.ifr_min),
        f.to_count,
        f.ldg_count,
        f.autoland,
        f.toga,
        round(f.pic_min),
        f.remarks,
    )


def read_filled_logbook_flights(path: Path) -> list[Flight]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    flights: list[Flight] = []
    for r in range(1, ws.max_row + 1):
        parsed = parse_date(ws.cell(r, 1).value)
        if not parsed:
            continue
        flt = norm_flight_number(ws.cell(r, 2).value)
        dep = norm_airport(ws.cell(r, 3).value)
        arr = norm_airport(ws.cell(r, 4).value)
        reg = norm_reg(ws.cell(r, 5).value)
        if not (flt and dep and arr and reg):
            continue
        flights.append(
            Flight(
                date=parsed,
                flt=flt,
                dep=dep,
                arr=arr,
                reg=reg,
                dc=str(ws.cell(r, 6).value or "").strip().upper(),
                off=ws.cell(r, 7).value if isinstance(ws.cell(r, 7).value, time) else None,
                on=ws.cell(r, 8).value if isinstance(ws.cell(r, 8).value, time) else None,
                blk_min=minutes(ws.cell(r, 9).value),
                takeoff=ws.cell(r, 10).value if isinstance(ws.cell(r, 10).value, time) else None,
                landing=ws.cell(r, 11).value if isinstance(ws.cell(r, 11).value, time) else None,
                act_min=minutes(ws.cell(r, 12).value),
                ngt_min=minutes(ws.cell(r, 13).value),
                ifr_min=minutes(ws.cell(r, 14).value),
                to_count=int(ws.cell(r, 15).value or 0),
                ldg_count=int(ws.cell(r, 16).value or 0),
                autoland=str(ws.cell(r, 17).value or ""),
                toga=str(ws.cell(r, 18).value or ""),
                pic_min=minutes(ws.cell(r, 19).value),
                remarks=str(ws.cell(r, 20).value or ""),
            )
        )
    return flights


def read_previous_output_flights() -> list[Flight]:
    return read_filled_logbook_flights(XLSX_OUT)


def compute_changes(previous: list[Flight], current: list[Flight]) -> tuple[int, int, int]:
    old = {flight_identity(f): flight_fingerprint(f) for f in previous}
    new = {flight_identity(f): flight_fingerprint(f) for f in current}
    added = len([key for key in new if key not in old])
    deleted = len([key for key in old if key not in new])
    modified = len([key for key in new if key in old and new[key] != old[key]])
    return added, deleted, modified


def norm_flight_number(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"(\d+)$", text)
    return match.group(1) if match else ""


def parse_date(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        nums = [int(x) for x in re.findall(r"\d+", value)]
        if len(nums) >= 3:
            return datetime(nums[0], nums[1], nums[2])
    return None


def configured_end_date() -> datetime | None:
    raw = os.environ.get("LOGBOOK_END_DATE", "").strip()
    return parse_date(raw) if raw else None


def filter_flights_by_end_date(flights: list[Flight]) -> list[Flight]:
    cutoff = configured_end_date()
    if not cutoff:
        return flights
    filtered = [flight for flight in flights if flight.date.date() <= cutoff.date()]
    print(f"LOGBOOK_END_DATE {cutoff.date()}")
    print(f"LOGBOOK_END_DATE_FILTERED {len(flights)}->{len(filtered)}")
    return filtered


def norm_airport(value) -> str:
    text = "" if value is None else str(value).strip().upper()
    return {"RKSI": "ICN", "RJAA": "NRT"}.get(text, text)


def norm_reg(value) -> str:
    text = "" if value is None else str(value).strip().upper()
    if text.endswith(".0"):
        text = text[:-2]
    return text[2:] if text.startswith("HL") else text


def page_chunks(flights: list[Flight]) -> list[list[Flight]]:
    return [flights[i : i + 10] for i in range(0, len(flights), 10)]


def compute_totals(pages: list[list[Flight]]) -> list[dict]:
    prev = START.copy()
    totals = []
    for page in pages:
        page_total = {
            "blk": round(sum(f.blk_min for f in page)),
            "ngt": round(sum(f.ngt_min for f in page)),
            "ifr": round(sum(f.ifr_min for f in page)),
            "to": sum(f.to_count for f in page),
            "ldg": sum(f.ldg_count for f in page),
            "pic": round(sum(f.pic_min for f in page)),
        }
        cumulative = {k: prev[k] + page_total[k] for k in prev}
        totals.append({"page": page_total, "previous": prev.copy(), "cumulative": cumulative.copy()})
        prev = cumulative
    return totals


def read_authoritative_totals() -> list[dict]:
    wb = load_workbook(PILOTLOG_EXPORT_XLSX, data_only=True)
    ws = wb.active
    totals = []
    r = 1
    while r <= ws.max_row:
        if ws.cell(r, 1).value == "DATE":
            page_row = r + 1
            while page_row <= ws.max_row and ws.cell(page_row, 1).value != "PAGE TOTAL":
                page_row += 1
            if page_row > ws.max_row:
                break
            cumulative_row = page_row + 2
            page = {
                "blk": minutes(ws.cell(page_row, 9).value),
                "ngt": minutes(ws.cell(page_row, 13).value),
                "ifr": minutes(ws.cell(page_row, 14).value),
                "to": int(ws.cell(page_row, 15).value or 0),
                "ldg": int(ws.cell(page_row, 16).value or 0),
                "pic": minutes(ws.cell(page_row, 19).value),
            }
            cumulative = {
                "blk": minutes(ws.cell(cumulative_row, 9).value),
                "ngt": minutes(ws.cell(cumulative_row, 13).value),
                "ifr": minutes(ws.cell(cumulative_row, 14).value),
                "to": int(ws.cell(cumulative_row, 15).value or 0),
                "ldg": int(ws.cell(cumulative_row, 16).value or 0),
                "pic": minutes(ws.cell(cumulative_row, 19).value),
            }
            previous = {
                "blk": cumulative["blk"] - page["blk"],
                "ngt": cumulative["ngt"] - page["ngt"],
                "ifr": cumulative["ifr"] - page["ifr"],
                "to": cumulative["to"] - page["to"],
                "ldg": cumulative["ldg"] - page["ldg"],
                "pic": cumulative["pic"] - page["pic"],
            }
            if not totals:
                previous = START.copy()
            totals.append({"page": page, "previous": previous, "cumulative": cumulative})
            r = page_row + 3
        r += 1
    return totals


def write_xlsx(flights: list[Flight], totals: list[dict]) -> None:
    wb = xlsxwriter.Workbook(XLSX_OUT)
    ws = wb.add_worksheet("flt_log")
    ws.set_landscape()
    ws.set_paper(9)
    ws.fit_to_pages(1, 0)
    ws.set_margins(0.25, 0.25, 0.35, 0.35)
    headers = ["DATE", "FLT", "FROM", "TO", "REG", "DC", "RO", "RI", "BLK", "TO", "LD", "ACT", "NGT", "INS", "이륙", "착륙", "자동착륙", "TOGA", "787 PIC", "Remarks"]
    header_fmt = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#1B3A6B", "border": 1, "border_color": "#B8962E", "align": "center", "valign": "vcenter", "font_size": 8})
    cell_fmt = wb.add_format({"border": 1, "border_color": "#BFC6D1", "align": "center", "valign": "vcenter", "font_size": 8})
    date_fmt = wb.add_format({"border": 1, "border_color": "#BFC6D1", "align": "center", "valign": "vcenter", "font_size": 8, "num_format": "yyyy-mm-dd"})
    time_fmt = wb.add_format({"border": 1, "border_color": "#BFC6D1", "align": "center", "valign": "vcenter", "font_size": 8, "num_format": "hh:mm"})
    dur_fmt = wb.add_format({"border": 1, "border_color": "#BFC6D1", "align": "center", "valign": "vcenter", "font_size": 8, "num_format": "[h]:mm"})
    total_lbl_fmt = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#1B3A6B", "border": 1, "border_color": "#B8962E", "align": "right", "valign": "vcenter", "font_size": 8})
    total_dur_fmt = wb.add_format({"bold": True, "bg_color": "#F4E8C4", "border": 1, "border_color": "#B8962E", "align": "center", "valign": "vcenter", "font_size": 8, "num_format": "[h]:mm"})
    total_int_fmt = wb.add_format({"bold": True, "bg_color": "#F4E8C4", "border": 1, "border_color": "#B8962E", "align": "center", "valign": "vcenter", "font_size": 8})
    for i, width in enumerate([11, 8, 7, 7, 7, 7, 6, 6, 7, 6, 6, 7, 7, 7, 5, 5, 8, 6, 8, 14]):
        ws.set_column(i, i, width)
    pages = page_chunks(flights)
    for pidx, page in enumerate(pages):
        start = pidx * 22
        for c, h in enumerate(headers):
            ws.write(start, c, h, header_fmt)
        for ridx in range(10):
            row = start + 1 + ridx
            excel_row = row + 1
            if ridx >= len(page):
                for c in range(20):
                    ws.write_blank(row, c, None, cell_fmt)
                continue
            f = page[ridx]
            ws.write_datetime(row, 0, f.date, date_fmt)
            for col, value in [(1, f.flt), (2, f.dep), (3, f.arr), (4, f.reg), (5, f.dc)]:
                ws.write(row, col, value, cell_fmt)
            if f.off:
                ws.write_number(row, 6, excel_time_fraction(f.off), time_fmt)
            else:
                ws.write_blank(row, 6, None, time_fmt)
            if f.on:
                ws.write_number(row, 7, excel_time_fraction(f.on), time_fmt)
            else:
                ws.write_blank(row, 7, None, time_fmt)
            if f.blk_min:
                ws.write_number(row, 8, duration_fraction(f.blk_min), dur_fmt)
            else:
                ws.write_blank(row, 8, None, dur_fmt)
            if f.takeoff:
                ws.write_number(row, 9, excel_time_fraction(f.takeoff), time_fmt)
            else:
                ws.write_blank(row, 9, None, time_fmt)
            if f.landing:
                ws.write_number(row, 10, excel_time_fraction(f.landing), time_fmt)
            else:
                ws.write_blank(row, 10, None, time_fmt)
            ws.write_formula(row, 11, f'=IF(J{excel_row}>0,MOD(K{excel_row}-J{excel_row},1),"")', dur_fmt, duration_fraction(f.act_min) if f.act_min else "")
            ws.write_number(row, 12, duration_fraction(f.ngt_min), dur_fmt)
            ws.write_number(row, 13, duration_fraction(f.ifr_min), dur_fmt)
            ws.write_number(row, 14, f.to_count, cell_fmt)
            ws.write_number(row, 15, f.ldg_count, cell_fmt)
            ws.write(row, 16, f.autoland, cell_fmt)
            ws.write(row, 17, f.toga, cell_fmt)
            parts = []
            for dc in ALL_DCS:
                factor = PIC_FACTORS[dc]
                result = f"I{excel_row}" if factor == 1.0 else (f"I{excel_row}*{factor:.12g}" if factor else '""')
                parts.append(f'F{excel_row}="{dc}",{result}')
            parts.append('TRUE,""')
            ws.write_formula(row, 18, "=_xlfn.IFS(" + ",".join(parts) + ")", dur_fmt, duration_fraction(f.pic_min) if f.pic_min else "")
            ws.write(row, 19, f.remarks, cell_fmt)
        page_total_row = start + 11
        prev_row = start + 12
        cum_row = start + 13
        for row, label in [(page_total_row, "PAGE TOTAL"), (prev_row, "PREVIOUS TOTAL"), (cum_row, "CUMULATIVE TOTAL (CHOI SANG JOON)")]:
            ws.merge_range(row, 0, row, 7, label, total_lbl_fmt)
            for c in range(8, 20):
                ws.write_blank(row, c, None, total_dur_fmt if c in [8, 11, 12, 13, 18] else total_int_fmt)
        first_data_row = start + 2
        last_data_row = start + 11
        for c, key in [(8, "blk"), (12, "ngt"), (13, "ifr"), (18, "pic")]:
            col = get_column_letter(c + 1)
            ws.write_formula(page_total_row, c, f"=SUM({col}{first_data_row}:{col}{last_data_row})", total_dur_fmt, duration_fraction(totals[pidx]["page"][key]))
        ws.write_formula(page_total_row, 11, f"=SUM(L{first_data_row}:L{last_data_row})", total_dur_fmt, duration_fraction(sum(f.act_min for f in page)))
        for c, key in [(14, "to"), (15, "ldg")]:
            col = get_column_letter(c + 1)
            ws.write_formula(page_total_row, c, f"=SUM({col}{first_data_row}:{col}{last_data_row})", total_int_fmt, totals[pidx]["page"][key])
        for row, name in [(prev_row, "previous"), (cum_row, "cumulative")]:
            for c, key in [(8, "blk"), (12, "ngt"), (13, "ifr"), (18, "pic")]:
                ws.write_number(row, c, duration_fraction(totals[pidx][name][key]), total_dur_fmt)
            for c, key in [(14, "to"), (15, "ldg")]:
                ws.write_number(row, c, totals[pidx][name][key], total_int_fmt)
    wb.close()


def draw_fit(c: canvas.Canvas, text: str, x: float, y: float, w: float, size: float = 6.0, bold: bool = False, align: str = "CENTER") -> None:
    text = "" if text is None else str(text)
    font = "Helvetica-Bold" if bold else "Helvetica"
    while size > 3.7 and stringWidth(text, font, size) > w - 2:
        size -= 0.2
    c.setFont(font, size)
    if align == "LEFT":
        c.drawString(x + 1.1, y, text)
    elif align == "RIGHT":
        c.drawRightString(x + w - 1.1, y, text)
    else:
        c.drawCentredString(x + w / 2, y, text)


def write_a4_pdf(flights: list[Flight], totals: list[dict]) -> None:
    pages = page_chunks(flights)
    c = canvas.Canvas(str(A4_OUT), pagesize=landscape(A4))
    width, height = landscape(A4)
    left, right = 8 * mm, 8 * mm
    table_w = width - left - right
    title = "PILOT FLIGHT LOGBOOK · CHOI SANG JOON · B787 Type Rating · ICAO Annex 1 / EASA FCL.050 Compliant Format"
    cols = [("DATE", 18), ("FLT", 13), ("FROM", 12), ("TO", 12), ("OFF BLOCK", 17), ("ON BLOCK", 17), ("TYPE(B787)", 15), ("REG(HL+REG)", 17), ("DC", 10), ("FLIGHT TIME", 18), ("NIGHT", 15), ("IFR", 15), ("T·O", 9), ("LDG", 9), ("AUTO LDG", 15), ("TOGA", 11), ("PIC TIME", 17), ("REMARKS", 24)]
    total_units = sum(w for _, w in cols)
    col_ws = [w / total_units * table_w for _, w in cols]
    row_h = 7.2 * mm
    head_h = 7 * mm
    for pidx, page in enumerate(pages):
        c.setFillColor(NAVY)
        c.rect(0, height - 17 * mm, width, 17 * mm, stroke=0, fill=1)
        c.setFillColor(GOLD)
        c.rect(0, height - 18 * mm, width, 1.2 * mm, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 10.4)
        c.drawCentredString(width / 2, height - 9.8 * mm, title)
        c.setFont("Helvetica", 7)
        c.drawRightString(width - right, height - 14.2 * mm, f"Page {pidx + 1} of {len(pages)}")
        y = height - 27 * mm
        x = left
        c.setFillColor(NAVY)
        c.rect(left, y - head_h, table_w, head_h, stroke=0, fill=1)
        for (label, _), cw in zip(cols, col_ws):
            c.setStrokeColor(GOLD)
            c.rect(x, y - head_h, cw, head_h, stroke=1, fill=0)
            c.setFillColor(colors.white)
            draw_fit(c, label, x, y - 4.8 * mm, cw, size=5.8, bold=True)
            x += cw
        y -= head_h
        for ridx in range(10):
            f = page[ridx] if ridx < len(page) else None
            values = [""] * len(cols)
            if f:
                values = [f.date.strftime("%Y-%m-%d"), f.flt, f.dep, f.arr, fmt_time(f.off), fmt_time(f.on), "B787", "HL" + f.reg if f.reg else "", f.dc, fmt_hmm(f.blk_min), fmt_hmm(f.ngt_min), fmt_hmm(f.ifr_min), str(f.to_count or ""), str(f.ldg_count or ""), f.autoland, f.toga, fmt_hmm(f.pic_min) if f.pic_min else "", pdf_remarks(f.remarks)]
            c.setFillColor(colors.HexColor("#F8FAFC") if ridx % 2 else colors.white)
            c.rect(left, y - row_h, table_w, row_h, stroke=0, fill=1)
            x = left
            for value, cw in zip(values, col_ws):
                c.setStrokeColor(GRID)
                c.rect(x, y - row_h, cw, row_h, stroke=1, fill=0)
                c.setFillColor(TEXT)
                draw_fit(c, value, x, y - 4.8 * mm, cw, size=5.8)
                x += cw
            y -= row_h
        label_w = sum(col_ws[:9])
        for label, total in [("PAGE TOTAL", totals[pidx]["page"]), ("PREVIOUS TOTAL", totals[pidx]["previous"]), ("CUMULATIVE TOTAL (CHOI SANG JOON)", totals[pidx]["cumulative"])]:
            c.setFillColor(NAVY)
            c.rect(left, y - row_h, label_w, row_h, stroke=0, fill=1)
            c.setStrokeColor(GOLD)
            c.rect(left, y - row_h, label_w, row_h, stroke=1, fill=0)
            c.setFillColor(colors.white)
            draw_fit(c, label, left, y - 4.8 * mm, label_w - 2, size=6.2, bold=True, align="RIGHT")
            x = left + label_w
            for idx in range(9, len(cols)):
                cw = col_ws[idx]
                c.setFillColor(LIGHT_GOLD)
                c.rect(x, y - row_h, cw, row_h, stroke=0, fill=1)
                c.setStrokeColor(GOLD)
                c.rect(x, y - row_h, cw, row_h, stroke=1, fill=0)
                value = ""
                if idx == 9:
                    value = fmt_hmm(total["blk"])
                elif idx == 10:
                    value = fmt_hmm(total["ngt"])
                elif idx == 11:
                    value = fmt_hmm(total["ifr"])
                elif idx == 12:
                    value = str(total["to"])
                elif idx == 13:
                    value = str(total["ldg"])
                elif idx == 16:
                    value = fmt_hmm(total["pic"])
                c.setFillColor(TEXT)
                draw_fit(c, value, x, y - 4.8 * mm, cw, size=6.0, bold=True)
                x += cw
            y -= row_h
        sig_y = 15 * mm
        c.setFillColor(TEXT)
        c.setFont("Helvetica", 7)
        c.drawString(left, sig_y + 6 * mm, "I certify that the entries on this page are true and correct.")
        c.line(width - 82 * mm, sig_y + 7 * mm, width - 10 * mm, sig_y + 7 * mm)
        c.drawCentredString(width - 46 * mm, sig_y + 3.3 * mm, "Pilot Signature")
        c.showPage()
    c.save()


def make_slot_page(path: Path, cover: bool) -> None:
    c = canvas.Canvas(str(path), pagesize=landscape(A4))
    w, h = landscape(A4)
    c.setFillColor(NAVY if cover else colors.white)
    c.rect(0, 0, w, h, stroke=0, fill=1)
    if cover:
        c.setFillColor(GOLD)
        c.rect(0, h - 18 * mm, w, 5 * mm, stroke=0, fill=1)
        c.rect(0, 13 * mm, w, 2 * mm, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredString(w / 2, h / 2 + 8 * mm, "PILOT FLIGHT LOGBOOK")
        c.setFillColor(LIGHT_GOLD)
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(w / 2, h / 2 - 6 * mm, "CHOI SANG JOON")
    else:
        c.setStrokeColor(colors.HexColor("#D7DCE3"))
        for y in [h - 25 * mm - i * 8 * mm for i in range(15)]:
            c.line(24 * mm, y, w - 24 * mm, y)
    c.save()


def write_booklet() -> None:
    reader = PdfReader(str(A4_OUT))
    cover_path = WORK / "booklet_cover_slot.pdf"
    back_path = WORK / "booklet_back_lines_slot.pdf"
    make_slot_page(cover_path, True)
    make_slot_page(back_path, False)
    back = PdfReader(str(back_path)).pages[0]
    content = [PdfReader(str(cover_path)).pages[0]] + [p for p in reader.pages] + [back]
    padded = math.ceil(len(content) / 4) * 4
    while len(content) < padded:
        content.insert(-1, copy(back))
    writer = PdfWriter()
    out_w, out_h = portrait(A4)
    slot_w, slot_h = landscape(A4)
    scale = min(out_w / slot_w, (out_h / 2) / slot_h)
    draw_w, draw_h = slot_w * scale, slot_h * scale
    xoff = (out_w - draw_w) / 2
    top_y = out_h / 2 + ((out_h / 2) - draw_h) / 2
    bot_y = ((out_h / 2) - draw_h) / 2
    for s in range(padded // 4):
        front_left = padded - 2 * s
        front_right = 2 * s + 1
        back_left = 2 * s + 2
        back_right = padded - 2 * s - 1
        for top_idx, bottom_idx in [(front_left, front_right), (back_left, back_right)]:
            sheet_writer = PdfWriter()
            sheet = sheet_writer.add_blank_page(width=out_w, height=out_h)
            sheet.merge_transformed_page(copy(content[top_idx - 1]), Transformation().scale(scale).translate(xoff, top_y), expand=False)
            sheet.merge_transformed_page(copy(content[bottom_idx - 1]), Transformation().scale(scale).translate(xoff, bot_y), expand=False)
            overlay = WORK / f"booklet_fold_{s}_{top_idx}_{bottom_idx}.pdf"
            oc = canvas.Canvas(str(overlay), pagesize=portrait(A4))
            oc.setStrokeColor(colors.HexColor("#9AA3AF"))
            oc.setDash(2, 3)
            oc.line(8 * mm, out_h / 2, out_w - 8 * mm, out_h / 2)
            oc.save()
            sheet.merge_page(PdfReader(str(overlay)).pages[0])
            writer.add_page(sheet)
    with open(BOOKLET_OUT, "wb") as fh:
        writer.write(fh)


def render_page(pdf: Path, page: int, name: str) -> Path | None:
    reader = PdfReader(str(pdf))
    if page < 1 or page > len(reader.pages):
        return None
    tmp = WORK / f"{name}_page_{page}.pdf"
    writer = PdfWriter()
    writer.add_page(reader.pages[page - 1])
    with open(tmp, "wb") as fh:
        writer.write(fh)
    try:
        subprocess.run(["qlmanage", "-t", "-s", "1800", "-o", str(WORK), str(tmp)], check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except FileNotFoundError:
        pass
    png = WORK / f"{tmp.name}.png"
    if png.exists():
        return png
    ppm_prefix = WORK / f"{tmp.stem}_poppler"
    try:
        subprocess.run(["pdftoppm", "-png", "-singlefile", "-r", "180", str(tmp), str(ppm_prefix)], check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except FileNotFoundError:
        pass
    poppler_png = ppm_prefix.with_suffix(".png")
    return poppler_png if poppler_png.exists() else None


def validate(flights: list[Flight], totals: list[dict]) -> dict:
    wb_formula = load_workbook(XLSX_OUT, data_only=False)
    ws_formula = wb_formula.active
    wb_values = load_workbook(XLSX_OUT, data_only=True)
    ws_values = wb_values.active
    formula_errors = 0
    bad_formats = []
    formulas = []
    for row in ws_formula.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(cell.value)
                if any(err in cell.value for err in ["#NAME?", "#VALUE!", "#REF!", "#DIV/0!", "#N/A"]):
                    formula_errors += 1
            if cell.column in [9, 13, 14, 19] and cell.value is not None and cell.row % 22 != 1:
                if cell.number_format != "[h]:mm":
                    bad_formats.append((cell.coordinate, cell.number_format))
    literal_errors = 0
    for row in ws_values.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and any(err in value for err in ["#NAME?", "#VALUE!", "#REF!", "#DIV/0!", "#N/A"]):
                literal_errors += 1
    if bad_formats:
        raise RuntimeError(f"Bad [h]:mm formats: {bad_formats[:8]}")
    if formula_errors or literal_errors:
        raise RuntimeError(f"Formula/literal errors: {formula_errors}/{literal_errors}")
    for pidx in [0, len(totals) // 2, len(totals) - 1]:
        start = pidx * 22
        for row_offset, total_name in [(12, "previous"), (13, "cumulative")]:
            row = start + row_offset + 1
            expected = totals[pidx][total_name]
            actual = {
                "blk": minutes(ws_values.cell(row, 9).value),
                "ngt": minutes(ws_values.cell(row, 13).value),
                "ifr": minutes(ws_values.cell(row, 14).value),
                "to": int(ws_values.cell(row, 15).value or 0),
                "ldg": int(ws_values.cell(row, 16).value or 0),
                "pic": minutes(ws_values.cell(row, 19).value),
            }
            if any(abs(actual[k] - expected[k]) > 0.01 for k in ["blk", "ngt", "ifr", "pic"]) or any(actual[k] != expected[k] for k in ["to", "ldg"]):
                raise RuntimeError(f"Total mismatch page {pidx + 1} {total_name}: {actual} != {expected}")
    a4_reader = PdfReader(str(A4_OUT))
    booklet_reader = PdfReader(str(BOOKLET_OUT))
    first_text = a4_reader.pages[0].extract_text() or ""
    last_text = a4_reader.pages[-1].extract_text() or ""
    final = totals[-1]["cumulative"]
    for expected in ["PAGE TOTAL", "PREVIOUS TOTAL", "CUMULATIVE TOTAL", "13135:59", fmt_hmm(final["blk"]), fmt_hmm(final["ngt"]), fmt_hmm(final["ifr"]), str(final["to"]), str(final["ldg"]), fmt_hmm(final["pic"])]:
        if expected not in first_text + last_text:
            raise RuntimeError(f"Missing PDF text: {expected}")
    renders = [
        render_page(A4_OUT, 1, "verify_a4"),
        render_page(A4_OUT, len(a4_reader.pages), "verify_a4"),
        render_page(BOOKLET_OUT, 1, "verify_booklet"),
        render_page(BOOKLET_OUT, max(1, len(booklet_reader.pages) // 2), "verify_booklet"),
    ]
    return {
        "flight_count": len(flights),
        "a4_pages": len(a4_reader.pages),
        "booklet_pages": len(booklet_reader.pages),
        "formula_count": len(formulas),
        "formula_errors": formula_errors + literal_errors,
        "final": final,
        "renders": [str(p) for p in renders if p],
    }


def main() -> None:
    OUT.mkdir(exist_ok=True)
    WORK.mkdir(exist_ok=True)
    previous = filter_flights_by_end_date(read_previous_output_flights())
    flights = filter_flights_by_end_date(read_authoritative_source_flights())
    added, deleted, modified = compute_changes(previous, flights)
    pages = page_chunks(flights)
    totals = compute_totals(pages)
    write_xlsx(flights, totals)
    write_a4_pdf(flights, totals)
    write_booklet()
    result = validate(flights, totals)
    final = result["final"]
    print("ADDED", added)
    print("DELETED", deleted)
    print("MODIFIED", modified)
    print("FINAL_FLIGHTS", result["flight_count"])
    print("FINAL_A4_PAGES", result["a4_pages"])
    print("BOOKLET_PAGES", result["booklet_pages"])
    print("FORMULA_COUNT", result["formula_count"])
    print("FORMULA_ERRORS", result["formula_errors"])
    print("FINAL_TOTALS", {k: fmt_hmm(v) if k in ["blk", "ngt", "ifr", "pic"] else v for k, v in final.items()})
    print("RENDERS", result["renders"])
    print("SOURCE", SYNCED_AUTHORITATIVE_XLSX if SYNCED_AUTHORITATIVE_XLSX.exists() else PILOTLOG_EXPORT_XLSX)
    print("XLSX", XLSX_OUT)
    print("A4", A4_OUT)
    print("BOOKLET", BOOKLET_OUT)


if __name__ == "__main__":
    main()
